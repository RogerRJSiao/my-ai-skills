#!/usr/bin/env python3
"""
inspect_submission.py — dump the raw facts needed to grade one S06 Excel
實作 submission against 評閱標準.xlsx.

This does NOT compute a score. It only extracts objective facts (pivot
table settings, chart references, freeze panes, merges, borders, filters,
formulas) so the grader (you, the model) can apply the 評閱標準 rules
without re-deriving XML-parsing logic from scratch every time.

Usage:
    python inspect_submission.py <submission.xlsx> [--question original.xlsx]

If --question is given, the script also reports whether 車輛銷售/武將能力's
raw A-column and row-2 header data was altered from the original.

Output: pretty-printed JSON on stdout.
"""
import sys
import json
import zipfile
import argparse
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required (preinstalled in the skill runtime).")

NS = {
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def read_zip_xml(zf, name):
    try:
        return ET.fromstring(zf.read(name))
    except KeyError:
        return None


def sheet_rels(zf, sheet_index):
    name = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
    try:
        root = ET.fromstring(zf.read(name))
    except KeyError:
        return {}
    out = {}
    for rel in root:
        out[rel.attrib["Id"]] = rel.attrib
    return out


def find_sheet_index(zf, sheet_name):
    """Map a worksheet display name (e.g. '薪資') to its sheetN.xml index."""
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    wb_rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {r.attrib["Id"]: r.attrib["Target"] for r in wb_rels_root}
    for sheet in wb_root.iter(f"{{{NS['main']}}}sheet"):
        if sheet.attrib.get("name") == sheet_name:
            rid = sheet.attrib[f"{{{NS['r']}}}id"]
            target = rid_to_target[rid]  # e.g. 'worksheets/sheet1.xml'
            idx = target.split("sheet")[-1].split(".xml")[0]
            return int(idx)
    return None


def pivot_info(zf, sheet_index):
    """Return pivot-table facts for the pivotTable attached to this sheet
    (via worksheet rels), or None if no pivot table is attached."""
    rels = sheet_rels(zf, sheet_index)
    pivot_target = None
    for rid, attrib in rels.items():
        if attrib["Type"].endswith("/pivotTable"):
            pivot_target = attrib["Target"].replace("../", "xl/")
            break
    if pivot_target is None:
        return None

    pt_root = ET.fromstring(zf.read(pivot_target))
    m = f"{{{NS['main']}}}"
    cache_id = pt_root.attrib.get("cacheId")

    # Resolve cacheId -> pivotCacheDefinitionN.xml via workbook.xml + its rels
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    wb_rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {r.attrib["Id"]: r.attrib["Target"] for r in wb_rels_root}
    cache_target = None
    for pc in wb_root.iter(f"{m}pivotCache"):
        if pc.attrib.get("cacheId") == cache_id:
            cache_target = "xl/" + rid_to_target[pc.attrib[f"{{{NS['r']}}}id"]]
            break

    cache_source_name = None
    field_shared_items = {}
    if cache_target:
        cache_root = ET.fromstring(zf.read(cache_target))
        src = cache_root.find(f"{m}cacheSource/{m}worksheetSource")
        if src is not None:
            cache_source_name = src.attrib.get("name") or src.attrib.get("ref")
        for i, cf in enumerate(cache_root.findall(f"{m}cacheFields/{m}cacheField")):
            items = [s.attrib.get("v") for s in cf.findall(f"{m}sharedItems/{m}s")]
            if items:
                field_shared_items[i] = items

    pivot_fields = pt_root.findall(f"{m}pivotFields/{m}pivotField")
    row_field_idx = [f.attrib.get("x") for f in pt_root.findall(f"{m}rowFields/{m}field")]
    col_field_idx = [f.attrib.get("x") for f in pt_root.findall(f"{m}colFields/{m}field")]

    def field_axis(idx):
        for i, f in enumerate(pivot_fields):
            if f.attrib.get("axis") in ("axisRow", "axisCol") and str(i):
                pass
        return None

    # visible items per axis field (col/row) — items without h="1" are shown
    def visible_labels(field_index):
        if field_index is None:
            return None
        pf = pivot_fields[int(field_index)]
        items = pf.findall(f"{m}items/{m}item")
        shared = field_shared_items.get(int(field_index), [])
        labels = []
        for it in items:
            if it.attrib.get("t") == "default":  # grand total placeholder
                continue
            if it.attrib.get("h") == "1":
                continue
            x = it.attrib.get("x")
            if x is None:
                x = "0"
            try:
                labels.append(shared[int(x)])
            except (IndexError, ValueError):
                labels.append(f"<item x={x}>")
        return labels

    data_fields = []
    for df in pt_root.findall(f"{m}dataFields/{m}dataField"):
        data_fields.append({
            "name": df.attrib.get("name"),
            "subtotal": df.attrib.get("subtotal"),
            "numFmtId": df.attrib.get("numFmtId"),
        })

    return {
        "pivot_target": pivot_target,
        "source_name_or_range": cache_source_name,
        "source_is_named_table": bool(
            cache_source_name and not any(c in cache_source_name for c in "$:")
        ),
        "row_field_index": row_field_idx,
        "col_field_index": col_field_idx,
        "row_visible_labels": visible_labels(row_field_idx[0]) if row_field_idx else None,
        "col_visible_labels": visible_labels(col_field_idx[0]) if col_field_idx else None,
        "data_fields": data_fields,
    }


def pivot_output_number_format(xlsx_path, sheet_name):
    """Read the actual displayed number_format of the pivot table's data
    cells (applyNumberFormats may be 0, so the cell's own style rules)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb[sheet_name]
    formats = set()
    # Pivot output typically starts a few columns right of the source table;
    # scan a generous window and collect formats on numeric cells.
    for row in ws.iter_rows(min_row=1, max_row=10, min_col=10, max_col=20):
        for c in row:
            if isinstance(c.value, (int, float)):
                formats.add(c.number_format)
    return sorted(formats)


def chart_info(zf, sheet_index):
    """Return facts about the first chart drawn on this sheet, or None."""
    rels = sheet_rels(zf, sheet_index)
    drawing_target = None
    for rid, attrib in rels.items():
        if attrib["Type"].endswith("/drawing"):
            drawing_target = attrib["Target"].replace("../", "xl/")
            break
    if drawing_target is None:
        return None
    drawing_rels_path = str(Path(drawing_target).parent / "_rels" / (Path(drawing_target).name + ".rels"))
    try:
        d_rels_root = ET.fromstring(zf.read(drawing_rels_path))
    except KeyError:
        return None
    chart_target = None
    for rel in d_rels_root:
        if rel.attrib["Type"].endswith("/chart"):
            chart_target = "xl/charts/" + rel.attrib["Target"].split("/")[-1]
            break
    if chart_target is None:
        return None

    c_root = ET.fromstring(zf.read(chart_target))
    c = f"{{{NS['c']}}}"

    def strcache_values(container):
        """container is a <c:cat> or <c:tx> element; drill into its
        strRef/strCache (dynamic) or strLit (literal) child to get values."""
        if container is None:
            return None
        cache = container.find(f"{c}strRef/{c}strCache")
        if cache is None:
            cache = container.find(f"{c}strLit")
        if cache is None:
            return None
        return [pt.find(f"{c}v").text for pt in cache.findall(f"{c}pt")]

    is_dynamic = c_root.find(f".//{c}numRef") is not None or c_root.find(f".//{c}strRef") is not None
    has_literal_only = (c_root.find(f".//{c}numLit") is not None or c_root.find(f".//{c}strLit") is not None) and not is_dynamic

    # chart type
    chart_type = None
    for tag in ("barChart", "radarChart", "lineChart", "pieChart"):
        if c_root.find(f".//{c}{tag}") is not None:
            chart_type = tag
            break

    # category axis values (first series' c:cat)
    first_cat = c_root.find(f".//{c}ser/{c}cat")
    categories = strcache_values(first_cat) if first_cat is not None else None

    # series/legend names (each ser's c:tx)
    series_names = []
    for ser in c_root.findall(f".//{c}ser"):
        tx = ser.find(f"{c}tx")
        if tx is not None:
            vals = strcache_values(tx)
            if vals:
                series_names.append(vals[0])

    # category cell reference (to detect data alteration later)
    cat_ref = None
    cat_strref = first_cat.find(f"{c}strRef/{c}f") if first_cat is not None else None
    if cat_strref is not None:
        cat_ref = cat_strref.text

    val_ax = c_root.find(f".//{c}valAx")
    y_max = y_major_unit = gridlines_none = None
    has_title = c_root.find(f".//{c}chart/{c}title") is not None
    auto_title_deleted = c_root.find(f".//{c}autoTitleDeleted")
    has_title = has_title and not (auto_title_deleted is not None and auto_title_deleted.attrib.get("val") == "1")

    if val_ax is not None:
        scaling = val_ax.find(f"{c}scaling")
        if scaling is not None:
            mx = scaling.find(f"{c}max")
            y_max = mx.attrib.get("val") if mx is not None else None
        mu = val_ax.find(f"{c}majorUnit")
        y_major_unit = mu.attrib.get("val") if mu is not None else None
        gl = val_ax.find(f"{c}majorGridlines")
        if gl is None:
            gridlines_none = True
        else:
            nofill = gl.find(f".//{{{NS['a']}}}noFill")
            gridlines_none = nofill is not None

    x_axis_title = c_root.find(f".//{c}catAx/{c}title") is not None
    y_axis_title = c_root.find(f".//{c}valAx/{c}title") is not None
    has_data_labels = False
    for dlbls in c_root.iter(f"{c}dLbls"):
        show_val = dlbls.find(f"{c}showVal")
        if show_val is not None and show_val.attrib.get("val") == "1":
            has_data_labels = True

    # radar-specific: marker symbol
    marker_symbols = [m.attrib.get("val") for m in c_root.iter(f"{c}symbol")]

    return {
        "chart_target": chart_target,
        "chart_type": chart_type,
        "is_dynamic": is_dynamic,
        "literal_only": has_literal_only,
        "category_ref": cat_ref,
        "categories": categories,
        "series_names": series_names,
        "y_max": y_max,
        "y_major_unit": y_major_unit,
        "gridlines_none": gridlines_none,
        "has_chart_title": has_title,
        "has_x_axis_title": x_axis_title,
        "has_y_axis_title": y_axis_title,
        "has_data_labels": has_data_labels,
        "marker_symbols": marker_symbols,
    }


def grades_sheet_info(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb["成績"]
    freeze = ws.freeze_panes
    merged = [str(r) for r in ws.merged_cells.ranges]
    a1_val = ws["A1"].value
    a1_align = ws["A1"].alignment.horizontal
    centered_across = a1_align in ("center", "centerContinuous")

    # locate the "班平均" row (usually 16, but check A column just in case)
    avg_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "班平均":
            avg_row = r
            break

    g_formulas_ok = all(
        isinstance(ws.cell(row=r, column=7).value, str) and
        "AVERAGE" in ws.cell(row=r, column=7).value.upper()
        for r in range(3, 16)
    )

    row_avg_formulas_ok = row_avg_integer_fmt = None
    g_avg_row_value = None
    if avg_row:
        row_avg_formulas_ok = all(
            isinstance(ws.cell(row=avg_row, column=col).value, str) and
            "AVERAGE" in ws.cell(row=avg_row, column=col).value.upper()
            for col in range(2, 7)  # B..F
        )
        row_avg_integer_fmt = all(
            ws.cell(row=avg_row, column=col).number_format in ("0", "0_)", "General")
            and ws.cell(row=avg_row, column=col).number_format != "General"
            for col in range(2, 7)
        )
        g_cell = ws.cell(row=avg_row, column=7)
        g_avg_row_value = {"value": g_cell.value, "is_empty": g_cell.value in (None, "")}

    # borders across A1:G{avg_row or 16}
    last_row = avg_row or 16
    missing_border = []
    for r in range(1, last_row + 1):
        for col in range(1, 8):
            b = ws.cell(row=r, column=col).border
            if not (b.top.style or b.bottom.style or b.left.style or b.right.style):
                missing_border.append(ws.cell(row=r, column=col).coordinate)

    # sort + filter (read the table definition directly, openpyxl doesn't expose customFilters)
    sort_desc = filter_gt70 = hidden_rows = None
    with zipfile.ZipFile(xlsx_path) as zf:
        for name in zf.namelist():
            if name.startswith("xl/tables/") and name.endswith(".xml"):
                root = ET.fromstring(zf.read(name))
                if root.attrib.get("name") == "表格2" or "個人平均" in ET.tostring(root, encoding="unicode"):
                    sort_desc = root.find(f".//{{{NS['main']}}}sortCondition") is not None
                    cf = root.find(f".//{{{NS['main']}}}customFilter")
                    filter_gt70 = cf is not None and cf.attrib.get("operator") == "greaterThan"
    hidden_rows = [r for r in range(1, ws.max_row + 1)
                   if ws.row_dimensions[r].hidden]

    return {
        "freeze_panes": freeze,
        "merged_ranges": merged,
        "a1_value": a1_val,
        "a1_center_or_merged": bool(merged) or centered_across,
        "header_text_preserved": a1_val == "經濟甲班 期中考成績表",
        "avg_row_index": avg_row,
        "g3_g15_average_formulas_ok": g_formulas_ok,
        "row_avg_formulas_ok": row_avg_formulas_ok,
        "row_avg_integer_format": row_avg_integer_fmt,
        "g_avg_row_cell": g_avg_row_value,
        "missing_border_cells": missing_border,
        "borders_complete": len(missing_border) == 0,
        "sorted_descending": sort_desc,
        "filter_gt70_applied": filter_gt70,
        "hidden_rows": hidden_rows,
    }


def raw_data_unchanged(xlsx_path, question_path, sheet_name, col_range, row2_range):
    """Compare A-column (categories) and row-2 (headers) between submission
    and the original question file, cell by cell."""
    wb_sub = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_q = openpyxl.load_workbook(question_path, data_only=True)
    ws_sub, ws_q = wb_sub[sheet_name], wb_q[sheet_name]
    diffs = []
    for coord in col_range + row2_range:
        v_sub, v_q = ws_sub[coord].value, ws_q[coord].value
        if v_sub != v_q:
            diffs.append({"cell": coord, "submission": v_sub, "question": v_q})
    return {"unchanged": len(diffs) == 0, "diffs": diffs}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("submission")
    ap.add_argument("--question", default=None,
                     help="path to the original S06_excel_question.xlsx, "
                          "used to detect altered raw data in 車輛銷售/武將能力")
    args = ap.parse_args()

    xlsx_path = args.submission
    out = {"file": xlsx_path}

    with zipfile.ZipFile(xlsx_path) as zf:
        salary_idx = find_sheet_index(zf, "薪資")
        vehicle_idx = find_sheet_index(zf, "車輛銷售")
        general_idx = find_sheet_index(zf, "武將能力")

        out["pivot_薪資"] = pivot_info(zf, salary_idx) if salary_idx else None
        out["chart_車輛銷售"] = chart_info(zf, vehicle_idx) if vehicle_idx else None
        out["chart_武將能力"] = chart_info(zf, general_idx) if general_idx else None

    if out["pivot_薪資"]:
        out["pivot_薪資"]["value_number_formats"] = pivot_output_number_format(xlsx_path, "薪資")

    out["成績"] = grades_sheet_info(xlsx_path)

    if args.question:
        out["車輛銷售_raw_data_check"] = raw_data_unchanged(
            xlsx_path, args.question, "車輛銷售",
            col_range=["A3", "A4", "A5", "A6"],
            row2_range=["B2", "C2", "D2"],
        )
        out["武將能力_raw_data_check"] = raw_data_unchanged(
            xlsx_path, args.question, "武將能力",
            col_range=["A3", "A4", "A5", "A6", "A7"],
            row2_range=["B2", "C2", "D2", "E2"],
        )

    print(json.dumps(out, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
