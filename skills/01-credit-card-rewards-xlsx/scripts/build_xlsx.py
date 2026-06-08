"""
build_xlsx.py
=============
輔助腳本：將已整理好的資料列（list of lists）輸出為標準格式的 xlsx。

使用方式（供 AI 呼叫）：
    from scripts.build_xlsx import write_xlsx
    write_xlsx(rows, output_path)

rows 格式（每列為 list，共 10 個欄位）：
    [信用卡名稱, 開始日, 結束日, 商家名稱, 回饋%數,
     方案切換, 相同商家, 最高回饋, 優惠備註, 查詢日期]
"""

from __future__ import annotations
import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 樣式常數 ──────────────────────────────────────────────────
HEADER_BG    = "1F3864"
ALT_ROW_BG   = "DCE6F1"
WHITE_ROW_BG = "FFFFFF"
EXIST_BG     = "FFF2CC"
EXIST_FG     = "7D6608"
MAX_BG       = "E2EFDA"
MAX_FG       = "375623"
FONT_NAME    = "Arial"

HEADERS = [
    "信用卡名稱", "開始日", "結束日", "商家名稱", "回饋%數",
    "方案切換", "相同商家", "最高回饋", "優惠備註", "查詢日期",
]
COL_WIDTHS  = [22, 14, 14, 32, 10, 16, 12, 12, 55, 14]
CENTER_COLS = {2, 3, 5, 6, 7, 8, 10}   # 1-indexed


# ── 分析函式 ──────────────────────────────────────────────────

def _normalize_merchant(name: str) -> str:
    """標準化商家名稱（僅用於比對，不寫入欄位）。"""
    if not name:
        return ""
    suffixes = [
        "(台新Pay)", "(LINE Pay)", "(Apple Pay)",
        "(慶生月)", "(童樂匯)", "(趣旅行)", "(集精選)", "(樂饗購)",
        "實體門市", "(海外)", "海外",
    ]
    n = str(name)
    for s in suffixes:
        n = n.replace(s, "")
    return n.strip().upper()


def _parse_pct(s) -> float:
    """把 '3.3%' 轉為 3.3；無法解析則回傳 0.0。"""
    if not s:
        return 0.0
    m = re.search(r"([\d.]+)%", str(s))
    return float(m.group(1)) if m else 0.0


def annotate_exist_max(rows: list[list]) -> list[list]:
    """
    填入相同商家（col 6，index 6）與最高回饋（col 7，index 7）。
    輸入 rows 的 index 6、7 可為空字串，本函式會覆寫。
    """
    keys = [_normalize_merchant(r[3]) for r in rows]
    counts = Counter(keys)

    # 每個標準化商家的最高回饋率
    max_pct: dict[str, float] = defaultdict(float)
    for key, row in zip(keys, rows):
        max_pct[key] = max(max_pct[key], _parse_pct(row[4]))

    result = []
    for key, row in zip(keys, rows):
        r = list(row)
        if counts[key] > 1:
            r[6] = "Exist"
            r[7] = "Max" if _parse_pct(r[4]) == max_pct[key] else ""
        else:
            r[6] = ""
            r[7] = ""
        result.append(r)
    return result


# ── 輸出函式 ──────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def write_xlsx(rows: list[list], output_path: str | Path) -> Path:
    """
    將資料列輸出為標準格式 xlsx。

    Parameters
    ----------
    rows        : 每列 10 個欄位的資料，不含標頭列。
                  index 6 / 7（相同商家 / 最高回饋）可為空，
                  本函式會自動執行 annotate_exist_max() 填入。
    output_path : 輸出檔案路徑（含檔名）。

    Returns
    -------
    Path : 實際輸出的檔案路徑。
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = annotate_exist_max(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "信用卡"

    # ── 標頭列 ──
    h_fill   = PatternFill("solid", start_color=HEADER_BG, end_color=HEADER_BG)
    h_font   = Font(bold=True, color="FFFFFF", name=FONT_NAME, size=11)
    h_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border   = _thin_border()

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font   = h_font
        cell.fill   = h_fill
        cell.alignment = h_align
        cell.border = border

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── 資料列 ──
    alt_fill   = PatternFill("solid", start_color=ALT_ROW_BG,   end_color=ALT_ROW_BG)
    white_fill = PatternFill("solid", start_color=WHITE_ROW_BG, end_color=WHITE_ROW_BG)
    exist_fill = PatternFill("solid", start_color=EXIST_BG,     end_color=EXIST_BG)
    max_fill   = PatternFill("solid", start_color=MAX_BG,       end_color=MAX_BG)

    for i, row_data in enumerate(rows):
        row_num   = i + 2
        base_fill = alt_fill if row_num % 2 == 0 else white_fill

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col in CENTER_COLS else "left",
                vertical="center",
                wrap_text=True,
            )

            if col == 7 and val == "Exist":
                cell.fill = exist_fill
                cell.font = Font(name=FONT_NAME, size=10, bold=True, color=EXIST_FG)
            elif col == 8 and val == "Max":
                cell.fill = max_fill
                cell.font = Font(name=FONT_NAME, size=10, bold=True, color=MAX_FG)
            else:
                cell.fill = base_fill
                cell.font = Font(name=FONT_NAME, size=10)

    wb.save(output_path)
    print(f"✅ 輸出完成：{output_path}（共 {len(rows)} 筆資料）")
    return output_path


# ── CLI 使用範例 ──────────────────────────────────────────────
if __name__ == "__main__":
    today = date.today().strftime("%Y-%m-%d")
    sample = [
        ["範例銀行 測試卡", today, "2026-12-31",
         "基本回饋-國內不限通路", "1%", "通用", "", "", "一般消費基本回饋", today],
        ["範例銀行 測試卡", today, "2026-12-31",
         "7-ELEVEN", "3%", "天天刷", "", "", "【天天刷3%】超商類別", today],
        ["其他銀行 A卡", today, "2026-12-31",
         "7-ELEVEN", "2%", "通用", "", "", "一般消費", today],
    ]
    write_xlsx(sample, f"outputs/信用卡回饋方案整理_{today}.xlsx")
